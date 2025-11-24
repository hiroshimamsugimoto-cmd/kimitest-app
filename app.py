import streamlit as st
import openpyxl
from datetime import datetime
from io import BytesIO
import base64
from decimal import Decimal, ROUND_HALF_UP
import json
import os
from openpyxl.styles import Alignment

# === 履歴ファイル ===
HISTORY_FILE = "history.json"

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return []
    return []

def save_history(history):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

history = load_history()

# === 設定 ===
TEMPLATE = "検査報告書_(株)広島フォーマット.xlsx"

# --- ExcelのROUND ---
def excel_round(value, digits):
    q = '1.' + '0' * digits
    return float(Decimal(str(value)).quantize(Decimal(q), rounding=ROUND_HALF_UP))

st.title("📘 気密試験記録 入力フォーム")

# --- 入力項目 ---
系統名 = st.text_input("系統名")
試験圧力 = st.text_input("試験圧力(MPa以上)", placeholder="例：0.8")
試験範囲 = st.text_input("試験範囲")
試験媒体 = st.text_input("試験媒体")

col_a, col_b = st.columns(2)
with col_a:
    放置時間値 = st.text_input("放置時間の値", placeholder="例：30")
with col_b:
    放置単位 = st.selectbox("単位を選択", ["min以上", "h以上"])

放置時間 = f"{放置時間値}{放置単位}" if 放置時間値 else ""

使用機器No = st.text_input("使用圧力計機器No.")
測定場所 = st.text_input("測定場所")

# --- 開始日時 ---
st.subheader("開始日時")
col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    開始日 = st.date_input("日付", key="start_date")
with col2:
    開始時 = st.text_input("時", value="", key="start_hour", placeholder="例：9")
with col3:
    開始分 = st.text_input("分", value="", key="start_minute", placeholder="例：30")

# --- 終了日時 ---
st.subheader("終了日時")
col4, col5, col6 = st.columns([2, 1, 1])
with col4:
    終了日 = st.date_input("日付", key="end_date")
with col5:
    終了時 = st.text_input("時", value="", key="end_hour", placeholder="例：10")
with col6:
    終了分 = st.text_input("分", value="", key="end_minute", placeholder="例：15")

# --- 測定値入力 ---
st.subheader("測定値入力")
col5, col6 = st.columns(2)
with col5:
    P1 = st.text_input("開始圧力 (MPa)", placeholder="例：0.8760")
with col6:
    T1 = st.text_input("開始温度 (℃)", placeholder="例：20.1")

col7, col8 = st.columns(2)
with col7:
    P2p = st.text_input("終了圧力 (MPa)", placeholder="例：0.8756")
with col8:
    T2 = st.text_input("終了温度 (℃)", placeholder="例：19.3")

# 担当者名のみ入力
試験実施者 = st.text_input("試験実施者（担当者名のみ）")

# --- 数値変換 ---
def safe_float(v):
    try:
        return float(v.strip()) if v else None
    except:
        return None

P1 = safe_float(P1)
T1 = safe_float(T1)
P2p = safe_float(P2p)
T2 = safe_float(T2)

# --- 判定・保存 ---
if st.button("判定・保存"):
    if None in (P1, T1, P2p, T2):
        st.warning("⚠ 圧力・温度のすべてを入力してください。")
    else:
        try:
            # --- 日時 ---
            try:
                開始日時 = datetime.combine(
                    開始日,
                    datetime.strptime(f"{int(開始時 or 0):02d}:{int(開始分 or 0):02d}", "%H:%M").time()
                )
                終了日時 = datetime.combine(
                    終了日,
                    datetime.strptime(f"{int(終了時 or 0):02d}:{int(終了分 or 0):02d}", "%H:%M").time()
                )
            except:
                開始日時 = datetime.combine(開始日, datetime.strptime("00:00", "%H:%M").time())
                終了日時 = datetime.combine(終了日, datetime.strptime("00:00", "%H:%M").time())

            # --- 補正後圧力（Excel 完全一致） ---
            P2_corr_raw = ((P1 + 0.1013) * (T2 + 273.15) / (T1 + 273.15)) - 0.1013
            P2_corr = float(Decimal(str(P2_corr_raw)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))

            # --- ΔP（Excel = E10 - J10） ---
            ΔP_raw = P2p - P2_corr
            ΔP = float(Decimal(str(ΔP_raw)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))

            # --- 判定範囲（Excel：開始圧力×1%） ---
            判定範囲_raw = P1 * 0.01
            判定範囲 = float(Decimal(str(判定範囲_raw)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))

            # --- 判定 ---
            合否 = "合格" if abs(ΔP) <= 判定範囲 else "不合格"
            色 = "green" if 合否 == "合格" else "red"

            # --- 結果表示 ---
            st.markdown("## 📊 計算結果")
            st.write(f"- 補正後終了圧力: **{P2_corr:.3f} MPa**")
            st.write(f"- 圧力変化量 ΔP: **{ΔP:.3f} MPa**")
            st.write(f"- 判定範囲: ±**{判定範囲:.3f} MPa**")
            st.markdown(f"### <span style='color:{色};'>判定: {合否}</span>", unsafe_allow_html=True)

            # --- 履歴保存 ---
            history.append({
                "日時": datetime.now().strftime("%Y/%m/%d %H:%M"),
                "測定場所": 測定場所,
                "系統名": 系統名,
                "試験圧(MPa以上)": 試験圧力,
                "P1": P1,
                "T1": T1,
                "P2p": P2p,
                "T2": T2,
                "P2補正": P2_corr,
                "ΔP": ΔP,
                "判定": 合否,
                "試験実施者": 試験実施者
            })
            save_history(history)

            # --- Excel 出力 ---
            wb = openpyxl.load_workbook(TEMPLATE)
            ws = wb["気密試験記録"]

            def write(ws, cell, value):
                try:
                    ws[cell].value = value
                except AttributeError:
                    r = ws[cell].row
                    c = ws[cell].column
                    ws.cell(row=r, column=c, value=value)

            # 基本項目
            write(ws, "D3", 系統名)
            write(ws, "D4", f"{試験圧力}MPa以上")
            write(ws, "M4", 試験範囲)
            write(ws, "D5", 試験媒体)
            write(ws, "M5", 放置時間)
            write(ws, "D6", 使用機器No)
            write(ws, "M6", 測定場所)
            write(ws, "D8", 開始日時.strftime("%Y/%m/%d %H:%M"))
            write(ws, "M8", 終了日時.strftime("%Y/%m/%d %H:%M"))

            write(ws, "A10", f"{P1:.4f}")
            write(ws, "C10", f"{T1:.1f}")
            write(ws, "E10", f"{P2p:.4f}")
            write(ws, "G10", f"{T2:.1f}")
            write(ws, "J10", f"{P2_corr:.3f}MPa")
            write(ws, "M10", f"{ΔP:.3f}MPa")
            write(ws, "O10", f"±{判定範囲:.3f}MPa")
            write(ws, "M11", 合否)

            # --- 実施者欄 2 行中央揃え ---
            COMPANY_NAME = "株式会社 広島"
            value = f"{COMPANY_NAME}\n{試験実施者}"
            write(ws, "E11", value)
            ws["E11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # --- Excel 保存 ---
            output = BytesIO()
            wb.save(output)
            excel_data = output.getvalue()
            filename = f"気密検査報告書_{datetime.now().strftime('%Y%m%d')}.xlsx"
            b64 = base64.b64encode(excel_data).decode()
            href = (
                f'<a href="data:application/vnd.openxmlformats-'
                f'officedocument.spreadsheetml.sheet;base64,{b64}" '
                f'download="{filename}">📥 Excelをダウンロード</a>'
            )
            st.markdown(href, unsafe_allow_html=True)

        except Exception as e:
            st.error(f"⚠ エラー: {e}")

# --- 履歴表示 ---
st.markdown("---")

def get_value(record, keys, default=""):
    for key in keys:
        if key in record:
            return record[key]
    return default

with st.expander("📱 過去の測定履歴", expanded=False):
    if len(history) == 0:
        st.info("まだ履歴がありません。")
    else:
        for record in reversed(history[-50:]):
            系統 = get_value(record, ["系統名"], "")
            測定場所_val = get_value(record, ["測定場所"], "")
            P1_val = get_value(record, ["P1"], "")
            T1_val = get_value(record, ["T1"], "")
            P2_val = get_value(record, ["P2p"], "")
            T2_val = get_value(record, ["T2"], "")
            ΔP値 = get_value(record, ["ΔP"], "")
            判定 = get_value(record, ["判定"], "")
            実施者 = get_value(record, ["試験実施者"], "")

            header_line = f"{record.get('日時','')}"
            if 実施者:
                header_line += f"　実施者：{実施者}"

            st.markdown(
                f"""
            <div style="padding:12px; margin:12px 0; border-radius:10px;
            border:1px solid #ddd; background:#fafafa;">
                {header_line}<br>
                測定場所：{測定場所_val}<br>
                系統名：{系統}<br>
                開始圧力：{P1_val} MPa　開始温度：{T1_val} ℃<br>
                終了圧力：{P2_val} MPa　終了温度：{T2_val} ℃<br>
                ΔP：{ΔP値} MPa<br>
                判定：{判定}
            </div>
            """,
                unsafe_allow_html=True
            )

# --- 履歴削除 ---
st.markdown("---")
st.markdown("### 🗑 履歴管理")

if st.button("⚠ 履歴をすべて削除する"):
    history.clear()
    save_history(history)
    st.success("履歴を削除しました。ページを再読み込みしてください。")
