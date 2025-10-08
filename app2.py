import streamlit as st
import openpyxl
from datetime import datetime
from io import BytesIO
import base64
import smtplib
from email.message import EmailMessage

# === 設定 ===
TEMPLATE = "気密試験記録.xlsx"  # 同じフォルダにテンプレートExcelを置く

st.title("📘 気密試験記録 入力フォーム")

# ======================================
# 📧 Outlook設定の永続保存
# ======================================

@st.cache_data
def save_credentials(email, password):
    """メール認証情報を保存"""
    return {"email": email, "password": password}

@st.cache_data
def load_credentials():
    """保存済みの認証情報を読み込み"""
    return {}

# --- Outlook設定入力欄 ---
st.subheader("📧 Outlook送信設定（初回のみ入力）")

stored = load_credentials()

email = st.text_input("自分のOutlookメールアドレス", value=stored.get("email", ""))
password = st.text_input("アプリパスワード（Outlook）", type="password", value=stored.get("password", ""))

if email and password:
    save_credentials(email, password)
    st.info("✅ メール設定を保存しました（ブラウザを閉じても保持されます）")

# ======================================
# 入力フォーム本体
# ======================================

系統名 = st.text_input("系統名")
試験圧力 = st.text_input("試験圧力 (MPa)")
試験範囲 = st.text_input("試験範囲")
試験媒体 = st.text_input("試験媒体")
放置時間 = st.text_input("放置時間 (h)", placeholder="例：10min以上")
使用機器No = st.text_input("使用圧力計機器No.")
測定場所 = st.text_input("測定場所")

# --- 開始日時 ---
st.subheader("開始日時")
col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    開始日 = st.date_input("日付", key="start_date")
with col2:
    開始時 = st.number_input("時", min_value=0, max_value=23, value=9, key="start_hour")
with col3:
    開始分 = st.number_input("分", min_value=0, max_value=59, value=0, key="start_minute")

# --- 終了日時 ---
st.subheader("終了日時")
col4, col5, col6 = st.columns([2, 1, 1])
with col4:
    終了日 = st.date_input("日付 ", key="end_date")
with col5:
    終了時 = st.number_input("時 ", min_value=0, max_value=23, value=10, key="end_hour")
with col6:
    終了分 = st.number_input("分 ", min_value=0, max_value=59, value=0, key="end_minute")

# --- 測定値入力 ---
st.subheader("測定値入力")
col7, col8 = st.columns(2)
with col7:
    P1 = st.text_input("開始圧力 (MPa)", placeholder="例：0.8760")
with col8:
    T1 = st.text_input("開始温度 (℃)", placeholder="例：20.1")

col9, col10 = st.columns(2)
with col9:
    P2p = st.text_input("終了圧力 (MPa)", placeholder="例：0.8756")
with col10:
    T2 = st.text_input("終了温度 (℃)", placeholder="例：19.3")

試験実施者 = st.text_input("試験実施者")

# --- 数値変換 ---
def safe_float(v):
    try:
        return float(v.strip()) if v else None
    except:
        return None

P1 = safe_float(P1)
T1 = safe_float(T1)
P2p = safe_float(P2p)
T2 = safe_float(T2)

# ======================================
# 🧮 メール送信関数
# ======================================
def send_mail_outlook(sender, password, receiver, filename, file_data):
    try:
        msg = EmailMessage()
        msg["Subject"] = "気密試験記録データ（自動送信）"
        msg["From"] = sender
        msg["To"] = receiver
        msg.set_content(
            "以下の試験記録を送信しました。\n\n"
            f"送信日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}\n"
            "添付ファイルをご確認ください。"
        )
        msg.add_attachment(
            file_data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

        with smtplib.SMTP("smtp.office365.com", 587) as smtp:
            smtp.starttls()
            smtp.login(sender, password)
            smtp.send_message(msg)

        return True
    except Exception as e:
        st.error(f"📨 メール送信エラー: {e}")
        return False

# ======================================
# 🧮 判定・保存ボタン
# ======================================

if st.button("判定・保存"):
    if None in (P1, T1, P2p, T2):
        st.warning("⚠ 圧力・温度のすべてを入力してください。")
    else:
        try:
            # --- 日時生成 ---
            開始日時 = datetime.combine(開始日, datetime.strptime(f"{開始時:02d}:{開始分:02d}", "%H:%M").time())
            終了日時 = datetime.combine(終了日, datetime.strptime(f"{終了時:02d}:{終了分:02d}", "%H:%M").time())

            # --- 補正計算 ---
            T1_K = T1 + 273.15
            T2_K = T2 + 273.15
            P2_corr = P2p * (T1_K / T2_K)
            ΔP = P2_corr - P1
            判定範囲 = P1 * 0.01  # ±1%

            合否 = "合格" if abs(ΔP) <= 判定範囲 else "不合格"
            色 = "green" if 合否 == "合格" else "red"

            # --- 結果表示 ---
            st.markdown("## 📊 計算結果（ボイル・シャルルの法則に基づく補正）")
            st.write(f"- 補正後終了圧力 P2_corr: **{P2_corr:.4f} MPa**")
            st.write(f"- 圧力変化量 ΔP: **{ΔP:.4f} MPa**")
            st.write(f"- 判定範囲: ±**{判定範囲:.4f} MPa**")
            st.markdown(f"### <span style='color:{色};'>判定結果: {合否}</span>", unsafe_allow_html=True)

            # --- Excel出力 ---
            wb = openpyxl.load_workbook(TEMPLATE)
            ws = wb["気密試験記録"]

            ws["D3"] = 系統名
            ws["D4"] = 試験圧力
            ws["M4"] = 試験範囲
            ws["D5"] = 試験媒体
            ws["M5"] = 放置時間
            ws["D6"] = 使用機器No
            ws["M6"] = 測定場所
            ws["D8"] = 開始日時.strftime("%Y/%m/%d %H:%M")
            ws["M8"] = 終了日時.strftime("%Y/%m/%d %H:%M")
            ws["A10"] = f"{P1:.4f}"
            ws["C10"] = f"{T1:.1f}"
            ws["E10"] = f"{P2p:.4f}"
            ws["G10"] = f"{T2:.1f}"
            ws["J10"] = f"{P2_corr:.4f}"
            ws["M10"] = f"{ΔP:.4f}"
            ws["O10"] = f"±{判定範囲:.4f}"
            ws["M11"] = 合否
            ws["E11"] = 試験実施者

            # --- 保存 ---
            output = BytesIO()
            wb.save(output)
            excel_data = output.getvalue()
            filename = f"気密試験記録_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # --- ダウンロードリンク ---
            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📥 Excelをダウンロード</a>'
            st.markdown(href, unsafe_allow_html=True)

            # --- Outlookメール送信 ---
            if email and password:
                st.info("📧 Outlookメールに送信中...")
                if send_mail_outlook(email, password, email, filename, excel_data):
                    st.success("✅ 自分のOutlookメールに送信しました！")
                else:
                    st.warning("⚠ 送信に失敗しました。アプリパスワードを確認してください。")
            else:
                st.warning("⚠ メール設定が未登録です。")

        except Exception as e:
            st.error(f"⚠ エラーが発生しました: {e}")
